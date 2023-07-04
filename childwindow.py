from PySide2.QtWidgets import *
from PySide2.QtUiTools import *
from PySide2.QtGui import *
from PySide2.QtCore import *
import FDMgroundwater.onedimensionflow as fo
import FDMgroundwater.twodimensionsflow as ft
import time
import numpy as np
import threading
import openpyxl
import psutil


class Set_fourier_series(QDialog):
    # 定义信号
    signal_fourier_series = Signal(int)

    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/Set_fourier_series.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 设置傅里叶级数
        self.fourier_series = 1000  # 默认为1000
        # 监测按钮《确认》
        self.ui.sure.clicked.connect(self.sure)
        # 监测按钮《返回》
        self.ui.back.clicked.connect(self.back)

    def sure(self):
        fourier_series = self.ui.spinBox_fourier_series.value()
        self.signal_fourier_series.emit(fourier_series)
        self.ui.close()

    def back(self):
        self.ui.close()


class Set_cpu_cores(QDialog):
    # 定义信号
    signal_cpu_cores = Signal(int)

    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/Set_cpu_cores.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 读取当前设备的CPU物理核心数目
        cpu_c = psutil.cpu_count(logical=False)
        self.ui.textBrowser.append('监测到当前设备CPU物理核心为' + str(cpu_c) + '个')
        # 监测按钮《确认》
        self.ui.sure.clicked.connect(self.sure)
        # 监测按钮《返回》
        self.ui.back.clicked.connect(self.back)

    def sure(self):
        cpu_cores = self.ui.verticalSlider.value()
        self.signal_cpu_cores.emit(cpu_cores)
        self.ui.close()

    def back(self):
        self.ui.close()


class Set_width(QDialog):
    # 定义信号
    signal_width = Signal(int)

    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/Set_width.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 设置含水层宽度
        self.B = 1  # 默认为1
        # 监测按钮《确认》
        self.ui.sure.clicked.connect(self.sure)
        # 监测按钮《返回》
        self.ui.back.clicked.connect(self.back)

    def sure(self):
        width = self.ui.spinBox_width.value()
        self.signal_width.emit(width)
        self.ui.close()

    def back(self):
        self.ui.close()


class About_this_program(QDialog):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/About_this_program.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《返回》
        self.ui.back.clicked.connect(self.back)

    def back(self):
        self.ui.close()


class One_dimension_confined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/oodcasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = fo.Confined_aquifer_SF()

    def flow_draw(self):
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.length(self.ui.length.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class One_dimension_confined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/oodcausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 数值解存放
        self.solve_fdm = None
        # 解析解存放
        self.solve_as = None
        # 傅里叶级数存放
        self.fourier_series = None
        # 解析解分配CPU核心数
        self.cpu_cores = None
        # 误差存放
        self.error = None
        # 相对误差存放
        self.relative_error = None
        # 均方误差存放
        self.mean_square_error = None
        # 空间相对差分步长
        self.relative_step_length = None
        # 时间相对差分步长
        self.relative_step_time = None
        # 压力扩散系数
        self.pressure_diffusion_coefficient = None
        # 数值解方法存放
        self.how_fdm = None
        # 选定时刻存放
        self.time_location = None
        # 程序计算状态
        self.solve_activity = False
        # 监测按钮《使用向后隐式差分计算数值解结果》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《使用Crank-Nicolson中心差分计算数值解结果》
        self.ui.solve_cn.clicked.connect(self.solve_cn)
        # 监测按钮《计算解析解》
        self.ui.solve_analytic_solution.clicked.connect(self.solve_analytic_solution)
        # 监测按钮《数值解表面图绘图》
        self.ui.draw_solve_surface.clicked.connect(self.draw_solve_surface)
        # 监测按钮《解析解表面图绘图》
        self.ui.draw_solve_analytic_solution_surface.clicked.connect(self.draw_solve_analytic_solution_surface)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《误差对比分析》
        self.ui.error_analysis.clicked.connect(self.error_analysis)
        # 监测按钮《绘制误差表面图》
        self.ui.draw_error.clicked.connect(self.draw_error)
        # 监测按钮《保存日志》
        self.ui.save_date.clicked.connect(self.save_date)
        # 监测按钮《绘制选定时刻的数值解线型图》
        self.ui.draw_solve_line.clicked.connect(self.draw_solve_line)
        # 监测按钮《绘制选定时刻的解析解线型图》
        self.ui.draw_solve_analytic_solution_line.clicked.connect(self.draw_solve_analytic_solution_line)
        # 监测按钮《绘制数值解和解析解对比图》
        self.ui.draw_complete.clicked.connect(self.draw_complete)
        # 监测按钮《进行水均衡计算》
        self.ui.hydrological_budget.clicked.connect(self.hydrological_budget)
        # 获取自编库中的类的用法
        self.flow = fo.Confined_aquifer_USF()
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))
        self.ui.textBrowser.append(self.time)
        wb = openpyxl.Workbook()
        ws1 = wb.create_sheet('info', 0)
        ws1.append(
            ['x轴轴长', '空间差分步长', '相对空间差分步长', 't轴轴长', '时间差分步长', '相对时间差分步长', '左边界',
             '右边界', '初始条件', '压力扩散系数', '傅里叶级数', '平均相对误差', '均方误差', '数值解差分格式'])
        wb.save('缓存/' + self.time + '承压含水层一维非稳定流data.xlsx')

    def set_time_choose_box(self):  # 设置时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time.setMaximum(time_all - 1)
        self.ui.textBrowser_time.setPlainText(
            '计算时长为：' + str(self.flow.tl) + '天，时间分割步长为：' + str(self.flow.st) + '天，共计有时刻' + str(
                time_all) + '个。')

    def set_time_start_choose_box(self):  # 设置水均衡开始时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time_start.setMaximum(time_all - 2)

    def set_time_end_choose_box(self):  # 设置水均衡结束时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time_end.setMinimum(1)
        self.ui.spinBox_time_end.setMaximum(time_all - 1)

    def solve(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在使用向后隐式差分进行数值解求解')
        self.how_fdm = '向后隐式差分'
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.T / self.flow.S))
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def solve_cn(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在使用Crank-Nicolson中心差分进行数值解求解')
        self.how_fdm = 'Crank-Nicolson中心差分'
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.T / self.flow.S))
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve_cn()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def draw_solve_surface(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        title = self.how_fdm + '数值解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(self.flow.st)
        self.flow.draw_surface(self.solve_fdm, title=title)

    def draw_solve_line(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = self.how_fdm + '数值解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.ui.textBrowser.append('第' + str(self.time_location) + '时刻数值解绘图，当前时刻各点解值为：')
        self.ui.textBrowser.append(str(self.solve_fdm[self.time_location]))
        self.flow.draw(self.solve_fdm, time=self.time_location, title=title)

    def draw_complete(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        # 判定能否绘图
        if self.solve_as is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = self.how_fdm + '数值解和傅里叶级数解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw_complete(self.solve_fdm, self.solve_as, time=self.time_location, title=title, label0='数值解',
                                label1='解析解')

    def solve_analytic_solution_threading(self):
        new_thread = threading.Thread(target=self.solve_analytic_solution())
        new_thread.start()
        new_thread.join()

    def solve_analytic_solution(self):
        # 判定能否使用解析解
        if self.ui.leakage_recharge.toPlainText() == '' or self.ui.leakage_recharge.toPlainText() == '0':
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '只有源汇项为0时才可以进行解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行解析解求解')
        self.ui.textBrowser.append(
            '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
                self.cpu_cores) + '个')
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.T / self.flow.S))
        start_time = time.perf_counter()
        self.solve_as = self.flow.solve_multi(fourier_series=self.fourier_series, cpu_cores=self.cpu_cores)
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def draw_solve_analytic_solution_surface(self):
        # 判定能否绘图
        if self.solve_as is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        title = '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
            self.cpu_cores) + '个'
        self.flow.draw_surface(self.solve_as, title=title)

    def draw_solve_analytic_solution_line(self):
        # 判定能否绘图
        if self.solve_as is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
            self.cpu_cores) + '个' + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.ui.textBrowser.append('第' + str(self.time_location) + '时刻解析解绘图，当前时刻各点解值为：')
        self.ui.textBrowser.append(str(self.solve_as[self.time_location]))
        self.flow.draw(self.solve_as, time=self.time_location, title=title)

    def return_main(self):
        self.ui.close()

    def error_analysis(self):
        # 分析校验
        if self.solve_as is not None or self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行解析解和数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        # X轴差分点的数目
        m = int(self.flow.xl / self.flow.sl) + 1
        # 时间轴差分点的数目
        n = int(self.flow.tl / self.flow.st) + 1
        self.error = np.zeros((n, m))
        error_all = 0
        error_out = 0
        relative_error = 0
        for k in range(0, n):  # 对时间进行扫描
            for i in range(0, m):  # 对空间进行扫描
                # 计算均方误差
                self.error[k, i] = (self.solve_fdm[k, i] - self.solve_as[k, i])
                error_all += self.error[k, i] * self.error[k, i]
                # 计算相对误差
                if self.solve_as[k, i] < 0.001:
                    error_out += 1
                else:
                    relative_error += abs(self.error[k, i]) / abs(self.solve_as[k, i])

        self.mean_square_error = error_all / (n * m)
        self.relative_error = relative_error / (n * m - error_out)
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('误差分析完毕:')
        self.ui.textBrowser.append('均方误差:' + str(self.mean_square_error))
        self.ui.textBrowser.append('平均相对误差:' + str(self.relative_error))
        self.ui.textBrowser.append('----------------')
        wb = openpyxl.load_workbook('缓存/' + self.time + '承压含水层一维非稳定流data.xlsx')
        ws = wb['info']
        ws.append(
            [self.flow.xl, self.flow.sl, self.relative_step_length, self.flow.tl, self.flow.st, self.relative_step_time,
             self.flow.h_l, self.flow.h_r, self.flow.ic, self.pressure_diffusion_coefficient, self.fourier_series,
             self.relative_error, self.mean_square_error, self.how_fdm])
        wb.save('缓存/' + self.time + '承压含水层一维非稳定流data.xlsx')

    def draw_error(self):
        # 判定能否绘图
        if self.solve_as is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行误差分析！')  # 未通过校验即报错
            return None  # 结束代码
        self.flow.draw_surface(self.error, title='绝对误差表面图')

    def save_date(self):
        # 获取当前系统时间戳
        t = time.localtime()
        filepath = QFileDialog.getExistingDirectory(self.ui, "选择文件存储路径")
        f_ = filepath + '/日志' + str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t)) + '.txt'
        f_xlsx = filepath + '/日志' + str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t)) + '承压含水层一维非稳定流.xlsx'
        file = open(f_, 'w')
        file.write(self.ui.textBrowser.toPlainText())

        file.close()
        wb = openpyxl.load_workbook('缓存/' + self.time + '承压含水层一维非稳定流data.xlsx')
        wb.save(f_xlsx)
        self.ui.textBrowser.append('日志已保存')

    def hydrological_budget(self):
        self.ui.textBrowser.append("水均衡计算结果：")
        t0 = self.ui.spinBox_time_start.value()
        t1 = self.ui.spinBox_time_end.value()
        self.ui.textBrowser.append('所选时间段：第' + str(t0 * self.flow.st) + '天至' + str(t1 * self.flow.st) + '天')
        self.ui.textBrowser.append('关于水均衡计算结果：负数是流入含水层，正数是流出含水层')
        if self.solve_fdm is not None:
            hb_list_fdm = self.flow.hydrological_budget(self.solve_fdm, t0, t1)
            self.ui.textBrowser.append(
                '数值解水均衡结果：左边界流量：' + str(hb_list_fdm[0]) + ' 右边界流量：' + str(hb_list_fdm[1]))
            self.ui.textBrowser.append('通过边界流量计算的含水层水量变化：' + str(hb_list_fdm[2]))
            self.ui.textBrowser.append('通过贮水系数计算的含水层水量的变化：' + str(hb_list_fdm[3]))
            self.ui.textBrowser.append('二者比值为：' + str(hb_list_fdm[4]))
        else:
            self.ui.textBrowser.append('没有数值解结果！')
        if self.solve_as is not None:
            hb_list_as = self.flow.hydrological_budget_analytic_solution(self.solve_as, t0, t1)
            self.ui.textBrowser.append(
                '解析解水均衡结果：左边界流量：' + str(hb_list_as[0]) + ' 右边界流量：' + str(hb_list_as[1]))
            self.ui.textBrowser.append('通过边界流量计算的含水层水量变化：' + str(hb_list_as[2]))
            self.ui.textBrowser.append('通过贮水系数计算的含水层水量的变化：' + str(hb_list_as[3]))
            self.ui.textBrowser.append('二者比值为：' + str(hb_list_as[4]))
        else:
            self.ui.textBrowser.append('没有解析解结果！')
        self.ui.textBrowser.append('')


class One_dimension_unconfined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ooduasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = fo.Unconfined_aquifer_SF()

    def flow_draw(self):
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.length(self.ui.length.toPlainText())
        self.flow.draw(self.flow.solve())

    def return_main(self):
        self.ui.close()


class One_dimension_unconfined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ooduausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 数值解存放
        self.solve_fdm = None
        # 解析解存放
        self.solve_as0 = None
        self.solve_as1 = None
        # 傅里叶级数存放
        self.fourier_series = None
        # 解析解分配CPU核心数
        self.cpu_cores = None
        # 误差存放
        self.error = None
        # 相对误差存放
        self.relative_error = None
        # 均方误差存放
        self.mean_square_error = None
        # 空间相对差分步长
        self.relative_step_length = None
        # 时间相对差分步长
        self.relative_step_time = None
        # 压力扩散系数
        self.pressure_diffusion_coefficient = None
        # 数值解方法存放
        self.how_fdm = None
        # 选定时刻存放
        self.time_location = None
        # 程序计算状态
        self.solve_activity = False
        # 监测按钮《计算数值解结果》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《数值解表面图绘图》
        self.ui.draw_solve_surface.clicked.connect(self.draw_solve_surface)
        # 监测按钮《计算参考厚度法的解析解结果》
        self.ui.solve_reference_thickness_method.clicked.connect(self.solve_reference_thickness_method)
        # 监测按钮《参考厚度解析解表面图绘图》
        self.ui.draw_solve_reference_thickness_method_surface.clicked.connect(
            self.draw_solve_reference_thickness_method_surface)
        # 监测按钮《计算平方法的解析解结果》
        self.ui.solve_square_method.clicked.connect(self.solve_square_method)
        # 监测按钮《平方法解析解表面图绘图》
        self.ui.draw_solve_square_method_surface.clicked.connect(self.draw_solve_square_method_surface)
        # 监测按钮《绘制选定时刻的数值解线型图》
        self.ui.draw_solve_line.clicked.connect(self.draw_solve_line)
        # 监测按钮《绘制选定时刻的参考厚度解析解线型图》
        self.ui.draw_solve_reference_thickness_line.clicked.connect(self.draw_solve_reference_thickness_line)
        # 监测按钮《绘制选定时刻的平方打解析解线型图》
        self.ui.draw_solve_square_line.clicked.connect(self.draw_solve_square_line)
        # 监测按钮《绘制数值解和解析解对比线型图》
        self.ui.draw_complete.clicked.connect(self.draw_complete)
        # 监测按钮《保存日志信息》
        self.ui.save_date.clicked.connect(self.save_date)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《进行水均衡计算》
        self.ui.hydrological_budget.clicked.connect(self.hydrological_budget)
        # 获取自编库中的类的用法
        self.flow = fo.Unconfined_aquifer_USF()
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))
        self.ui.textBrowser.append(self.time)

    def set_time_choose_box(self):  # 设置时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time.setMaximum(time_all - 1)
        self.ui.textBrowser_time.setPlainText(
            '计算时长为：' + str(self.flow.tl) + '天，时间分割步长为：' + str(self.flow.st) + '天，共计有时刻' + str(
                time_all) + '个。')

    def set_time_start_choose_box(self):  # 设置水均衡开始时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time_start.setMaximum(time_all - 2)

    def set_time_end_choose_box(self):  # 设置水均衡结束时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time_end.setMinimum(1)
        self.ui.spinBox_time_end.setMaximum(time_all - 1)

    def solve(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.flow.specific_yield(self.ui.specific_yield.toPlainText())
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.how_fdm = '向后隐式差分'
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.K / self.flow.Sy))
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def solve_reference_thickness_method(self):
        # 判定能否使用解析解
        if self.ui.leakage_recharge.toPlainText() == '' or self.ui.leakage_recharge.toPlainText() == '0':
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '只有源汇项为0时才可以进行解析解计算！')  # 未通过校验即报错
            return 0  # 结束代码
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
        self.flow.specific_yield(self.ui.specific_yield.toPlainText())
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在使用参考厚度法进行解析解求解')
        self.ui.textBrowser.append(
            '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
                self.cpu_cores) + '个')
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.K * self.flow.ha / self.flow.Sy))
        start_time = time.perf_counter()
        if self.ui.checkBox.isChecked():
            self.solve_as0 = self.flow.solve_reference_thickness_method(fourier_series=self.fourier_series)
        else:
            self.solve_as0 = self.flow.solve_reference_thickness_method_multi(fourier_series=self.fourier_series,
                                                                              cpu_cores=self.cpu_cores)
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def solve_square_method(self):
        # 判定能否使用解析解
        if self.ui.leakage_recharge.toPlainText() == '' or self.ui.leakage_recharge.toPlainText() == '0':
            self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '只有源汇项为0时才可以进行解析解计算！')  # 未通过校验即报错
            return 0  # 结束代码
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.reference_thickness(self.ui.reference_thickness.toPlainText())
        self.flow.specific_yield(self.ui.specific_yield.toPlainText())
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 相对空间差分步长和相对时间差分步长的设定
        self.relative_step_length = self.flow.sl / self.flow.xl
        self.relative_step_time = self.flow.st / self.flow.tl
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在使用平方法进行解析解求解')
        self.ui.textBrowser.append(
            '傅里叶级数解（傅里叶级数取前' + str(self.fourier_series) + '项)，' + '分配CPU核心' + str(
                self.cpu_cores) + '个')
        self.ui.textBrowser.append(
            '相对空间差分步长：' + str(self.relative_step_length) + '相对时间差分步长：' + str(self.relative_step_time))
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        self.ui.textBrowser.append('压力扩散系数:' + str(self.flow.K * self.flow.ha / self.flow.Sy))
        start_time = time.perf_counter()
        if self.ui.checkBox.isChecked():
            self.solve_as1 = self.flow.solve_square_method(fourier_series=self.fourier_series)
        else:
            self.solve_as1 = self.flow.solve_square_method_multi(fourier_series=self.fourier_series,
                                                                 cpu_cores=self.cpu_cores)
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()
        self.set_time_start_choose_box()
        self.set_time_end_choose_box()

    def draw_solve_surface(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        title = '数值解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(self.flow.st)
        self.flow.draw_surface(self.solve_fdm, title=title)

    def draw_solve_reference_thickness_method_surface(self):
        # 判定能否绘图
        if self.solve_as0 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行参考厚度法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        title = '参考厚度法解析解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(self.flow.st)
        self.flow.draw_surface(self.solve_as0, title=title)

    def draw_solve_square_method_surface(self):
        # 判定能否绘图
        if self.solve_as1 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行平方法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        title = '平方法解析解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(self.flow.st)
        self.flow.draw_surface(self.solve_as1, title=title)

    def draw_solve_line(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '数值解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw(self.solve_fdm, time=self.time_location, title=title)
        self.ui.textBrowser.append('第' + str(self.time_location) + '时刻数值解绘图，当前时刻各点解值为：')
        self.ui.textBrowser.append(str(self.solve_fdm[self.time_location]))

    def draw_solve_reference_thickness_line(self):
        # 判定能否绘图
        if self.solve_as0 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行参考厚度法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '参考厚度法解析解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw(self.solve_as0, time=self.time_location, title=title)
        self.ui.textBrowser.append('第' + str(self.time_location) + '时刻参考厚度法解析解解绘图，当前时刻各点解值为：')
        self.ui.textBrowser.append(str(self.solve_as0[self.time_location]))

    def draw_solve_square_line(self):
        # 判定能否绘图
        if self.solve_as1 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行平方法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '平方法解析解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw(self.solve_as1, time=self.time_location, title=title)
        self.ui.textBrowser.append('第' + str(self.time_location) + '时刻平方法解析解绘图，当前时刻各点解值为：')
        self.ui.textBrowser.append(str(self.solve_as1[self.time_location]))

    def draw_complete(self):
        # 判定能否绘图
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
            # 判定能否绘图
        if self.solve_as0 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行参考厚度法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        # 判定能否绘图
        if self.solve_as0 is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行平方法解析解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '数值解和傅里叶级数解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw_complete(self.solve_fdm, self.solve_as0, self.solve_as1, time=self.time_location, title=title,
                                label0='数值解',
                                label1='参考厚度法解析解', label2='平方法解析解')

    def save_date(self):
        # 获取当前系统时间戳
        t = time.localtime()
        filepath = QFileDialog.getExistingDirectory(self.ui, "选择文件存储路径")
        f_ = filepath + '/日志' + str(time.strftime("%Y-%m-%d_%H时%M分%S秒，潜水含水层一维非稳定流", t)) + '.txt'
        file = open(f_, 'w')
        file.write(self.ui.textBrowser.toPlainText())
        file.close()

    def hydrological_budget(self):
        self.ui.textBrowser.append("水均衡计算结果：")
        t0 = self.ui.spinBox_time_start.value()
        t1 = self.ui.spinBox_time_end.value()
        self.ui.textBrowser.append('所选时间段：第' + str(t0 * self.flow.st) + '天至' + str(t1 * self.flow.st) + '天')
        self.ui.textBrowser.append('关于水均衡计算结果：负数是流入含水层，正数是流出含水层')
        if self.solve_fdm is not None:
            hb_list_fdm = self.flow.hydrological_budget(self.solve_fdm, t0, t1)
            self.ui.textBrowser.append(
                '数值解水均衡结果：左边界流量：' + str(hb_list_fdm[0]) + ' 右边界流量：' + str(hb_list_fdm[1]))
            self.ui.textBrowser.append('通过边界流量计算的含水层水量变化：' + str(hb_list_fdm[2]))
            self.ui.textBrowser.append('通过给水度计算的含水层水量的变化：' + str(hb_list_fdm[3]))
            self.ui.textBrowser.append('二者比值为：' + str(hb_list_fdm[4]))
        else:
            self.ui.textBrowser.append('没有数值解结果！')
        if self.solve_as0 is not None:
            hb_list_as = self.flow.hydrological_budget_analytic_solution(self.solve_as0, t0, t1)
            self.ui.textBrowser.append(
                '参考厚度法解析解水均衡结果：左边界流量：' + str(hb_list_as[0]) + ' 右边界流量：' + str(hb_list_as[1]))
            self.ui.textBrowser.append('通过边界流量计算的含水层水量变化：' + str(hb_list_as[2]))
            self.ui.textBrowser.append('通过给水度计算的含水层水量的变化：' + str(hb_list_as[3]))
            self.ui.textBrowser.append('二者比值为：' + str(hb_list_as[4]))
        else:
            self.ui.textBrowser.append('没有参考厚度法解析解结果！')
        if self.solve_as1 is not None:
            hb_list_as = self.flow.hydrological_budget_analytic_solution(self.solve_as1, t0, t1)
            self.ui.textBrowser.append(
                '平方法解析解水均衡结果：左边界流量：' + str(hb_list_as[0]) + ' 右边界流量：' + str(hb_list_as[1]))
            self.ui.textBrowser.append('通过边界流量计算的含水层水量变化：' + str(hb_list_as[2]))
            self.ui.textBrowser.append('通过给水度计算的含水层水量的变化：' + str(hb_list_as[3]))
            self.ui.textBrowser.append('二者比值为：' + str(hb_list_as[4]))
        else:
            self.ui.textBrowser.append('没有平方法解析解结果！')
        self.ui.textBrowser.append(' ')

    def return_main(self):
        self.ui.close()


class Two_dimension_confined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ttdcasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 预设差分结果矩阵
        self.solve_fdm = None
        # 监测按钮《数值解计算》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Confined_aquifer_SF()
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))

    def solve(self):
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.ui.textBrowser.append('空间差分步长：' + str(self.flow.sl))
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')

    def flow_draw(self):
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.flow.draw(self.solve_fdm)

    def return_main(self):
        self.ui.close()


class Two_dimension_unconfined_aquifer_stable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ttduasf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 预设差分结果矩阵
        self.solve_fdm = None
        # 监测按钮《数值解计算》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Unconfined_aquifer_SF()
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))

    def solve(self):
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.ui.textBrowser.append('空间差分步长：' + str(self.flow.sl))
        start_time = time.perf_counter()
        self.solve_fdm = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')

    def flow_draw(self):
        if self.solve_fdm is not None:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.flow.draw(self.solve_fdm)

    def return_main(self):
        self.ui.close()


class Two_dimension_confined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ttdcausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 预设差分结果矩阵
        self.solve_fdm = None
        # 监测按钮《数值解计算》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 监测按钮《上一时刻》
        self.ui.previous_time.clicked.connect(self.previous_time)
        # 监测按钮《下一时刻》
        self.ui.next_time.clicked.connect(self.next_time)
        # 获取自编库中的类的用法
        self.flow = ft.Confined_aquifer_USF()
        # 存储水头解值的列表
        self.h_all_time = []
        self.time_location = 0  # 时刻位置
        self.time_all = 0  # 抛开初始时刻的所有时刻个数
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))

    def solve(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.flow.storativity(self.ui.storativity.toPlainText())
        self.flow.transmissivity(self.ui.transmissivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.ui.textBrowser.append('空间差分步长：' + str(self.flow.sl))
        start_time = time.perf_counter()
        self.h_all_time = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')

    def flow_draw(self):
        if self.h_all_time:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.flow.draw(self.h_all_time[0])  # 绘制初始时刻的水头值
        self.time_all = len(self.h_all_time) - 1
        self.time_location = 0
        self.ui.progressBar.reset()
        self.ui.progressBar.setValue(0)  # 进度条置为0

    def next_time(self):
        self.time_location += 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def previous_time(self):
        self.time_location -= 1
        self.flow.draw(self.h_all_time[self.time_location])
        self.ui.progressBar.reset()
        T = int((self.time_location / self.time_all) * 100)
        self.ui.progressBar.setValue(T)  # 设置进度条进度

    def return_main(self):
        self.ui.close()


class Two_dimension_unconfined_aquifer_unstable_flow(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/ttduausf.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《计算并绘图》
        self.ui.solve.clicked.connect(self.solve)
        # 监测按钮《绘制选定时刻的数值解表面图》
        self.ui.draw_solve_surface.clicked.connect(self.draw_solve_surface)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Unconfined_aquifer_USF()
        # 存储水头解值的列表
        self.h_all_time = []
        self.time_location = 0  # 时刻位置
        self.time_all = 0  # 抛开初始时刻的所有时刻个数
        # 获取当前系统时间戳
        t = time.localtime()
        # 日志时间
        self.time = str(time.strftime("%Y-%m-%d_%H时%M分%S秒", t))
        self.ui.textBrowser.append('程序运行时间（北京时间）：' + str(time.strftime("%Y-%m-%d %H:%M:%S", t)))

    def set_time_choose_box(self):  # 设置时刻选择条
        time_all = int(self.flow.tl / self.flow.st) + 1
        self.ui.spinBox_time.setMaximum(time_all - 1)
        self.ui.textBrowser_time.setPlainText(
            '计算时长为：' + str(self.flow.tl) + '天，时间分割步长为：' + str(self.flow.st) + '天，共计有时刻' + str(
                time_all) + '个。')

    def solve(self):
        self.flow.l_boundary(self.ui.l_boundary.toPlainText())
        self.flow.r_boundary(self.ui.r_boundary.toPlainText())
        self.flow.t_boundary(self.ui.t_boundary.toPlainText())
        self.flow.b_boundary(self.ui.b_boundary.toPlainText())
        self.flow.step_length(self.ui.step_length.toPlainText())
        self.flow.step_time(self.ui.step_time.toPlainText())
        self.flow.x_length(self.ui.x_length.toPlainText())
        self.flow.y_length(self.ui.y_length.toPlainText())
        self.flow.t_length(self.ui.t_length.toPlainText())
        self.flow.initial_condition(self.ui.initial_condition.toPlainText())
        self.flow.specific_yield(self.ui.specific_yield.toPlainText())
        self.flow.hydraulic_conductivity(self.ui.hydraulic_conductivity.toPlainText())
        self.flow.leakage_recharge(self.ui.leakage_recharge.toPlainText())
        # 获取当前系统时间戳
        t = time.localtime()
        self.ui.textBrowser.append(str(time.strftime("%H:%M:%S", t)))
        self.ui.textBrowser.append('正在进行数值解求解')
        self.ui.textBrowser.append(
            '空间差分步长：' + str(self.flow.sl) + ' 时间差分步长：' + str(self.flow.st))
        start_time = time.perf_counter()
        self.h_all_time = self.flow.solve()
        end_time = time.perf_counter()
        self.ui.textBrowser.append('计算完毕，用时' + str(end_time - start_time) + '秒')
        # 设置时刻选择条
        self.set_time_choose_box()

    def draw_solve_surface(self):
        if self.h_all_time:
            pass
        else:
            QMessageBox.critical(self.ui, '错误', '请先进行数值解计算！')  # 未通过校验即报错
            return None  # 结束代码
        self.time_location = self.ui.spinBox_time.value()
        title = '数值解，空间差分步长为' + str(self.flow.sl) + '时间差分步长为' + str(
            self.flow.st) + '，绘图时刻为第' + str(self.time_location) + '时刻'
        self.flow.draw(self.h_all_time[self.time_location], title=title)

    def return_main(self):
        self.ui.close()


class Two_dimension_Toth_difficult_baisn(QMainWindow):
    def __init__(self):
        super().__init__()
        # 从文件中加载ui格式
        self.ui = QUiLoader().load("ui/toth.ui")
        self.ui.setWindowIcon(QIcon("water.ico"))
        # 监测按钮《多年平均水位绘图》
        self.ui.draw.clicked.connect(self.flow_draw)
        # 监测按钮《返回上一级》
        self.ui.back.clicked.connect(self.return_main)
        # 获取自编库中的类的用法
        self.flow = ft.Toth_difficult_baisn()

    def flow_draw(self):
        self.flow.basin_length(self.ui.basin_length.toPlainText())
        self.flow.basin_high(self.ui.basin_high.toPlainText())
        self.flow.average_water_level_equation(self.ui.average_water_level_equation.toPlainText())
        self.flow.draw_water_level()

    def return_main(self):
        self.ui.close()
